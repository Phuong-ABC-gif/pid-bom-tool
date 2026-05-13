"""
bom_generator.py — Xuất BOM đúng chuẩn ANHMINH FORM_BOM_STANDARD
Cấu trúc: Header công ty (row 1-15) + Header cột (row 16-18) + Data (row 19+) + Footer
Màu: Major=Cyan/Đỏ, Sub=Vàng/Đỏ, Data=Trắng/Đen
"""

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from pid_spec import get_line_category

# ── Hằng số màu & style ────────────────────────────────────────────────────────
BG_CYAN    = PatternFill("solid", start_color="00FFFF", end_color="00FFFF")
BG_YELLOW  = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")
BG_NONE    = PatternFill(fill_type=None)
FG_RED     = "FF0000"
FG_BLACK   = "000000"
FG_BLUE    = "000080"

COL_WIDTHS = {  # A..J
    "A": 5.57, "B": 47.71, "C": 15.14, "D": 16.14,
    "E": 11.29, "F": 12.00, "G": 10.71, "H": 15.43,
    "I": 6.71,  "J": 9.71,
}

COLS = list("ABCDEFGHIJ")


def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, value, font_color=FG_BLACK, bold=False,
          fill=BG_NONE, align="center", size=11, wrap=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, size=size, color=font_color)
    c.fill      = fill
    c.border    = _thin()
    c.alignment = Alignment(
        horizontal="left" if (align == "left" or col == 2) else "center",
        vertical="center", wrap_text=wrap
    )
    return c


def _row_h(ws, row, h=15):
    ws.row_dimensions[row].height = h


# ── Header công ty ─────────────────────────────────────────────────────────────

def _write_company_header(ws, project: str, doc_no: str, today: str):
    ws.merge_cells("A1:J1")
    ws["A1"].value = "ANHMINH"
    ws["A1"].font  = Font(name="Arial", bold=True, size=24, color=FG_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    _row_h(ws, 1, 30)

    ws.merge_cells("A2:J2")
    ws["A2"].value = "TECHNOLOGY TRADING COMPANY LIMITED"
    ws["A2"].font  = Font(name="Arial", bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    _row_h(ws, 2, 18)

    texts = [
        "Address: Lot B3, Street 9, Tan Tao Industrial Park, Binh Tan Dist., HCM City",
        "Tel: +84-28-37650639   Fax: +84-28-37650640",
        "Email: info@anhminhtech.com   Tax Code: 0304973037",
    ]
    for i, txt in enumerate(texts, 3):
        ws.merge_cells(f"A{i}:J{i}")
        ws[f"A{i}"].value = txt
        ws[f"A{i}"].font  = Font(name="Arial", size=11)
        ws[f"A{i}"].alignment = Alignment(horizontal="center", vertical="center")
        _row_h(ws, i, 15)

    ws.merge_cells("A6:J6")
    ws["A6"].value = "FACSIMILE TRANSMISSION"
    ws["A6"].font  = Font(name="Arial", bold=True, size=14)
    ws["A6"].alignment = Alignment(horizontal="center", vertical="center")
    _row_h(ws, 6, 20)

    info = [
        ("Date / Ngày:", today, "BG No. / BG số:", doc_no),
        ("To / Gởi đến:", "", "No. of pages / Số trang:", ""),
        ("From / Gởi từ:", "Anh Minh Tech", "Project / Dự án:", project),
        ("Subject / Tiêu đề:", "QUOTATION / BÁO GIÁ", "", ""),
    ]
    for i, row_data in enumerate(info, 7):
        ws.merge_cells(f"A{7+i-7}:B{7+i-7}") if False else None
        # Left: label + value
        ws.cell(row=7+i-7, column=1, value=row_data[0]).font = Font(name="Arial", bold=True, size=11)
        ws.cell(row=7+i-7, column=2, value=row_data[1]).font = Font(name="Arial", size=11)
        if row_data[2]:
            ws.cell(row=7+i-7, column=6, value=row_data[2]).font = Font(name="Arial", bold=True, size=11)
            ws.cell(row=7+i-7, column=7, value=row_data[3]).font = Font(name="Arial", size=11)
        _row_h(ws, 7+i-7, 15)

    for r in range(7, 12):
        _row_h(ws, r, 15)

    ws.merge_cells("A12:J12")
    ws["A12"].value = "BÁO GIÁ / QUOTATION"
    ws["A12"].font  = Font(name="Arial", bold=True, size=18, color=FG_RED)
    ws["A12"].alignment = Alignment(horizontal="center", vertical="center")
    _row_h(ws, 12, 28)

    ws.merge_cells("A13:J13")
    ws["A13"].value = "Kính gửi Quý Khách hàng, chúng tôi xin trân trọng báo giá như sau:"
    ws["A13"].font  = Font(name="Arial", size=11, italic=True)
    ws["A13"].alignment = Alignment(horizontal="left", vertical="center")
    _row_h(ws, 13, 15)

    for r, txt in [(14, "A / PHẠM VI CÔNG VIỆC / SCOPE OF WORK"), (15, "B / ĐƠN GIÁ / PRICE")]:
        ws.cell(row=r, column=1, value=txt).font = Font(name="Arial", bold=True, size=11)
        _row_h(ws, r, 15)


# ── Header cột ─────────────────────────────────────────────────────────────────

COL_HEADERS = ["STT", "MÔ TẢ", "CHỦNG LOẠI", "VẬT LIỆU",
               "K.THƯỚC 1", "K.THƯỚC 2", "TIÊU CHUẨN", "XUẤT XỨ", "ĐƠN VỊ", "SỐ LƯỢNG"]


def _write_col_header(ws, start_row=16):
    ws.merge_cells(f"{start_row}:{start_row+2}")
    for c_idx, (col, hdr) in enumerate(zip(COLS, COL_HEADERS), 1):
        ws.merge_cells(f"{col}{start_row}:{col}{start_row+2}")
        cell = ws.cell(row=start_row, column=c_idx, value=hdr)
        cell.font      = Font(name="Arial", bold=True, size=11, color=FG_RED)
        cell.fill      = BG_CYAN
        cell.border    = _thin()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(start_row, start_row + 3):
        _row_h(ws, r, 20)


# ── Section writers ────────────────────────────────────────────────────────────

def _write_major(ws, row, stt, label):
    vals = [stt, label] + [""] * 8
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font      = Font(name="Arial", bold=True, size=11, color=FG_RED)
        cell.fill      = BG_CYAN
        cell.border    = _thin()
        cell.alignment = Alignment(
            horizontal="left" if c == 2 else "center",
            vertical="center", wrap_text=True
        )
    _row_h(ws, row, 15)


def _write_sub(ws, row, stt, label):
    vals = [stt, label] + [""] * 8
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font      = Font(name="Arial", bold=True, size=11, color=FG_RED)
        cell.fill      = BG_YELLOW
        cell.border    = _thin()
        cell.alignment = Alignment(
            horizontal="left" if c == 2 else "center",
            vertical="center", wrap_text=True
        )
    _row_h(ws, row, 15)


def _write_item(ws, row, stt, mo_ta, chung_loai, vat_lieu,
                kt1, kt2, tieu_chuan, xuat_xu, dv, sl):
    vals = [stt, mo_ta, chung_loai, vat_lieu, kt1, kt2,
            tieu_chuan, xuat_xu, dv, sl]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font      = Font(name="Arial", bold=False, size=11, color=FG_BLACK)
        cell.fill      = BG_NONE
        cell.border    = _thin()
        cell.alignment = Alignment(
            horizontal="left" if c == 2 else "center",
            vertical="center", wrap_text=True
        )
    _row_h(ws, row, 15)


# ── Footer ─────────────────────────────────────────────────────────────────────

FOOTER_LINES = [
    ("TOTAL (VND)", True),
    ("VAT 10% (VND)", True),
    ("GRAND TOTAL (VND)", True),
    ("* Ghi chú / Notes:", True),
    ("- Giá trên chưa bao gồm chi phí vận chuyển, lắp đặt và thuế VAT.", False),
    ("- Không bao gồm chi phí bảo hiểm và thông quan.", False),
    ("C  ĐIỀU KIỆN BÁN HÀNG / SALES CONDITIONS:", True),
    ("- Địa điểm giao hàng / Delivery destination: Tại kho ANHMINH", False),
    ("- Thời gian giao hàng / Lead time: 7–10 tuần kể từ ngày xác nhận đơn hàng", False),
    ("- Điều kiện thanh toán / Payment term: 30% tạm ứng – 40% tập kết – 30% sau hoàn thành", False),
    ("- Thời hạn báo giá / Validity until: 30 ngày kể từ ngày gửi", False),
    ("Chân thành cảm ơn Quý Khách đã tin tưởng ANHMINH!", False),
    ("Trân trọng kính chào / Best regards,", False),
    ("ANHMINH TECHNOLOGY TRADING CO., LTD", True),
]


def _write_footer(ws, row):
    for label, bold in FOOTER_LINES:
        ws.merge_cells(f"A{row}:J{row}")
        c = ws[f"A{row}"]
        c.value     = label
        c.font      = Font(name="Arial", bold=bold, size=11,
                           color=FG_RED if bold and label.startswith(("TOTAL", "VAT", "GRAND", "*", "C ")) else FG_BLACK)
        c.alignment = Alignment(horizontal="center" if bold and "VND" in label else "left",
                                vertical="center", wrap_text=True)
        c.border    = _thin()
        _row_h(ws, row, 15)
        row += 1
    return row


# ── Main API ───────────────────────────────────────────────────────────────────

LINE_ORDER = ["SANITARY", "ICE_WATER", "COOLING", "STEAM", "AIR", "UNKNOWN"]

SUB_LABELS = {
    "SANITARY":  "ĐƯỜNG PRODUCT / CIP / PROCESS WATER (VI SINH)",
    "ICE_WATER": "ĐƯỜNG CHILLER (ICE WATER – CÔNG NGHIỆP)",
    "COOLING":   "ĐƯỜNG COOLING WATER (CÔNG NGHIỆP)",
    "STEAM":     "ĐƯỜNG STEAM / CONDENSATE (CÔNG NGHIỆP)",
    "AIR":       "ĐƯỜNG AIR – KHÍ NÉN (CÔNG NGHIỆP)",
    "UNKNOWN":   "CHƯA XÁC ĐỊNH LINE",
}

ALPHA = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")
ROMAN = ["I","II","III","IV","V","VI","VII","VIII","IX","X"]


def generate_bom(bom_rows: list[dict],
                 project: str = "",
                 doc_no: str = "",
                 warnings: list[str] | None = None) -> bytes:
    """
    Trả về bytes của file .xlsx (BOM chuẩn ANHMINH).
    bom_rows: output của dxf_reader.aggregate_bom()
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"

    # Column widths
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    today = date.today().strftime("%d/%m/%Y")
    _write_company_header(ws, project or "P&ID BOM", doc_no or "-", today)
    _write_col_header(ws, start_row=16)

    r = 19
    _write_major(ws, r, "A", "PHẦN CƠ KHÍ"); r += 1

    # Group by line_type
    from itertools import groupby
    by_type: dict[str, list] = {lt: [] for lt in LINE_ORDER}
    for row in bom_rows:
        lt = row.get("line_type", "UNKNOWN")
        by_type.setdefault(lt, []).append(row)

    sub_idx = 0
    item_idx = 0

    for lt in LINE_ORDER:
        items = by_type.get(lt, [])
        if not items:
            continue

        sub_label = SUB_LABELS.get(lt, lt)
        _write_sub(ws, r, ROMAN[sub_idx % len(ROMAN)], sub_label)
        r += 1
        sub_idx += 1
        item_idx = 0

        # Sort: by block_name, then size
        items.sort(key=lambda x: (x.get("block_name",""), x.get("size","?")))

        for row_data in items:
            item_idx += 1
            stt = str(item_idx)
            mo_ta = row_data.get("block_name", "")
            cl    = row_data.get("chung_loai", "-")
            vl    = row_data.get("vat_lieu",   "-")
            kt1   = row_data.get("size",        "?")
            kt2   = ""
            tc    = row_data.get("tieu_chuan",  "-")
            xu    = ""
            dv    = row_data.get("don_vi",      "pcs")
            sl    = row_data.get("sl",           1)

            # Thêm ghi chú Thread end nếu áp dụng
            if row_data.get("conn_type") == "Thread end" and lt != "SANITARY":
                mo_ta = f"{mo_ta} (Thread end)"

            _write_item(ws, r, stt, mo_ta, cl, vl, kt1, kt2, tc, xu, dv, sl)
            r += 1

    # Cảnh báo
    if warnings:
        _write_sub(ws, r, "⚠", "CẢNH BÁO — CẦN KIỂM TRA THỦ CÔNG"); r += 1
        for i, w in enumerate(warnings, 1):
            _write_item(ws, r, str(i), w, "", "", "", "", "", "", "", ""); r += 1

    _write_footer(ws, r)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
