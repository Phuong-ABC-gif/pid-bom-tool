"""
PID BOM Extractor — SPX / Tetra Pak
Anh Minh 2021 Standard
Rule-based, no AI cost, reads block names from DXF.
"""

import io
import math
import re
import tempfile
from collections import defaultdict

import ezdxf
import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ═══════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════

IGNORE_BLOCKS = {
    "flow arrow", "linenametpleft", "linenametpright",
    "pid", "objects", "title block", "revision", "border",
    "north arrow", "scale bar", "flow_arrow",
}

MAX_DIST = 15  # DXF units — proximity radius for size detection

# ───── COLOUR CONSTANTS ─────
BG_CYAN   = PatternFill("solid", start_color="00FFFF", end_color="00FFFF")
BG_YELLOW = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")
BG_WHITE  = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
RED       = "FF0000"
BLACK     = "000000"
BLUE      = "003399"


# ═══════════════════════════════════════════════════════════════════
#  EQUIPMENT LIBRARY  (block-name → display info)
# ═══════════════════════════════════════════════════════════════════

# fmt: off
# Each entry:
#   key   = canonical name (lowercase, words only — used for matching)
#   value = dict with MO_TA, CHUNG_LOAI, VAT_LIEU, TIEU_CHUAN, DON_VI, SECTION
#           SECTION: "sanitary" | "ice_water" | "cooling" | "steam" | "air" | "instrument" | "equipment"

SPX_LIBRARY = {
    # ── SINGLE SEAT VALVE ──────────────────────────────────────────
    "ssv 200 sv41 l manual":           {"mo_ta":"Single Seat Valve 200 (SV41-L) – Manual",         "chung_loai":"Manual",     "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 200 sv41 l actuator":         {"mo_ta":"Single Seat Valve 200 (SV41-L) – Actuator NC",    "chung_loai":"Actuator NC","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 200 sv41 l thinktop":         {"mo_ta":"Single Seat Valve 200 (SV41-L) – Thinktop",       "chung_loai":"Thinktop",   "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 200 nc sw41 l thinktop":      {"mo_ta":"Single Seat Valve 200 (SV41-L) – Thinktop NC",    "chung_loai":"Thinktop",   "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 200 nc sw41 l actuator":      {"mo_ta":"Single Seat Valve 200 (SV41-L) – Actuator NC",    "chung_loai":"Actuator NC","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 200 manual":                  {"mo_ta":"Single Seat Valve 200 – Manual",                  "chung_loai":"Manual",     "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 200 actuator":                {"mo_ta":"Single Seat Valve 200 – Actuator NC",              "chung_loai":"Actuator NC","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 200 thinktop":                {"mo_ta":"Single Seat Valve 200 – Thinktop",                 "chung_loai":"Thinktop",   "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 200 nc air actuator":         {"mo_ta":"Single Seat Valve 200 – Air Actuator NC",          "chung_loai":"Actuator NC","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 200 no air actuator":         {"mo_ta":"Single Seat Valve 200 – Air Actuator NO",          "chung_loai":"Actuator NO","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},

    "ssv 210 sv43 ll manual":          {"mo_ta":"Single Seat Valve 210 (SV43-LL) – Manual",        "chung_loai":"Manual",     "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 210 sv43 ll actuator":        {"mo_ta":"Single Seat Valve 210 (SV43-LL) – Actuator NC",   "chung_loai":"Actuator NC","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 210 sv43 ll thinktop":        {"mo_ta":"Single Seat Valve 210 (SV43-LL) – Thinktop",      "chung_loai":"Thinktop",   "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 210 manual":                  {"mo_ta":"Single Seat Valve 210 – Manual",                  "chung_loai":"Manual",     "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 210 actuator":                {"mo_ta":"Single Seat Valve 210 – Actuator NC",              "chung_loai":"Actuator NC","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 210 thinktop":                {"mo_ta":"Single Seat Valve 210 – Thinktop",                 "chung_loai":"Thinktop",   "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 210 nc air actuator":         {"mo_ta":"Single Seat Valve 210 – Air Actuator NC",          "chung_loai":"Actuator NC","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 210 no air actuator":         {"mo_ta":"Single Seat Valve 210 – Air Actuator NO",          "chung_loai":"Actuator NO","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 210 no n air actuator":       {"mo_ta":"Single Seat Valve 210 – Air Actuator NO-N",        "chung_loai":"Actuator NO","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},

    "ssv 220 sv44 tl manual":          {"mo_ta":"Single Seat Valve 220 (SV44-TL) – Manual",        "chung_loai":"Manual",     "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 220 sv44 tl actuator":        {"mo_ta":"Single Seat Valve 220 (SV44-TL) – Actuator NC",   "chung_loai":"Actuator NC","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 220 sv44 tl thinktop":        {"mo_ta":"Single Seat Valve 220 (SV44-TL) – Thinktop",      "chung_loai":"Thinktop",   "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 220 manual":                  {"mo_ta":"Single Seat Valve 220 – Manual",                  "chung_loai":"Manual",     "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 220 actuator":                {"mo_ta":"Single Seat Valve 220 – Actuator NC",              "chung_loai":"Actuator NC","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 220 thinktop":                {"mo_ta":"Single Seat Valve 220 – Thinktop",                 "chung_loai":"Thinktop",   "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},

    "ssv 300 sv42 t manual":           {"mo_ta":"Single Seat Valve 300 (SV42-T) – Manual",         "chung_loai":"Manual",     "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 300 sv42 t actuator":         {"mo_ta":"Single Seat Valve 300 (SV42-T) – Actuator NC",    "chung_loai":"Actuator NC","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 300 sv42 t thinktop":         {"mo_ta":"Single Seat Valve 300 (SV42-T) – Thinktop",       "chung_loai":"Thinktop",   "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 300 manual":                  {"mo_ta":"Single Seat Valve 300 – Manual",                  "chung_loai":"Manual",     "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 300 actuator":                {"mo_ta":"Single Seat Valve 300 – Actuator NC",              "chung_loai":"Actuator NC","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 300 thinktop":                {"mo_ta":"Single Seat Valve 300 – Thinktop",                 "chung_loai":"Thinktop",   "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},

    # ── MIXPROOF VALVE ─────────────────────────────────────────────
    "mixproof valve actuator":         {"mo_ta":"Mixproof Valve – Actuator NC",                    "chung_loai":"Actuator NC","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "mixproof valve thinktop":         {"mo_ta":"Mixproof Valve – Thinktop",                       "chung_loai":"Thinktop",   "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "mixproof valve nc solenoid":      {"mo_ta":"Mixproof Valve NC – Solenoid",                    "chung_loai":"Solenoid",   "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "mpv actuator":                    {"mo_ta":"Mixproof Valve – Actuator NC",                    "chung_loai":"Actuator NC","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "mpv thinktop":                    {"mo_ta":"Mixproof Valve – Thinktop",                       "chung_loai":"Thinktop",   "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},

    # ── BUTTERFLY VALVE SANITARY ───────────────────────────────────
    "butterfly valve manual":          {"mo_ta":"Butterfly Valve – Manual",                        "chung_loai":"Manual",     "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "butterfly valve actuator":        {"mo_ta":"Butterfly Valve – Actuator NC",                   "chung_loai":"Actuator NC","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "butterfly valve thinktop":        {"mo_ta":"Butterfly Valve – Thinktop",                      "chung_loai":"Thinktop",   "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "butterfly valve nc air actuator": {"mo_ta":"Butterfly Valve NC – Air Actuator",               "chung_loai":"Actuator NC","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "butterfly valve no air actuator": {"mo_ta":"Butterfly Valve NO – Air Actuator",               "chung_loai":"Actuator NO","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "bfv manual":                      {"mo_ta":"Butterfly Valve – Manual",                        "chung_loai":"Manual",     "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "bfv actuator":                    {"mo_ta":"Butterfly Valve – Actuator NC",                   "chung_loai":"Actuator NC","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "bfv thinktop":                    {"mo_ta":"Butterfly Valve – Thinktop",                      "chung_loai":"Thinktop",   "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "bfv tk":                          {"mo_ta":"Butterfly Valve – Thinktop",                      "chung_loai":"Thinktop",   "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},

    # ── SANITARY COMPONENTS ────────────────────────────────────────
    "non return valve":                {"mo_ta":"Non-Return Valve",                                "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "nrv":                             {"mo_ta":"Non-Return Valve",                                "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "sight glass":                     {"mo_ta":"Sight Glass",                                     "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"set","section":"sanitary"},
    "sg":                              {"mo_ta":"Sight Glass",                                     "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"set","section":"sanitary"},
    "sampling valve actuator":         {"mo_ta":"Sampling Valve – Actuator",                       "chung_loai":"Actuator",   "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "sampling valve manual":           {"mo_ta":"Sampling Valve – Manual",                         "chung_loai":"Manual",     "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "sva":                             {"mo_ta":"Sampling Valve – Actuator",                       "chung_loai":"Actuator",   "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "svm":                             {"mo_ta":"Sampling Valve – Manual",                         "chung_loai":"Manual",     "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "constant pressure valve":         {"mo_ta":"Constant Pressure Valve",                         "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "cpv":                             {"mo_ta":"Constant Pressure Valve",                         "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "sterile filter":                  {"mo_ta":"Sterile Filter",                                  "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "sf":                              {"mo_ta":"Sterile Filter",                                  "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "inline filter fsh":               {"mo_ta":"Inline Filter FSH",                               "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "inline filter fsr":               {"mo_ta":"Inline Filter FSR",                               "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "fsh":                             {"mo_ta":"Inline Filter FSH",                               "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "fsr":                             {"mo_ta":"Inline Filter FSR",                               "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "plate heat exchanger":            {"mo_ta":"Plate Heat Exchanger",                            "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"set","section":"sanitary"},
    "phe":                             {"mo_ta":"Plate Heat Exchanger",                            "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"set","section":"sanitary"},
    "tubular heat exchanger":          {"mo_ta":"Tubular Heat Exchanger",                          "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"set","section":"sanitary"},
    "the":                             {"mo_ta":"Tubular Heat Exchanger",                          "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"set","section":"sanitary"},
    "flexible hose":                   {"mo_ta":"Flexible Hose",                                   "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "fh":                              {"mo_ta":"Flexible Hose",                                   "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},

    # ── PUMPS SANITARY ─────────────────────────────────────────────
    "centrifugal pump w+":             {"mo_ta":"Centrifugal Pump W+",                             "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "centrifugal pump ws+":            {"mo_ta":"Centrifugal Pump WS+",                            "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "cp w+":                           {"mo_ta":"Centrifugal Pump W+",                             "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "cp ws+":                          {"mo_ta":"Centrifugal Pump WS+",                            "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "lobe pump dw+":                   {"mo_ta":"Lobe Pump DW+",                                   "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "lp dw+":                          {"mo_ta":"Lobe Pump DW+",                                   "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "peristaltic pump":                {"mo_ta":"Peristaltic Pump",                                "chung_loai":"-",          "vat_lieu":"-",     "tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "prp":                             {"mo_ta":"Peristaltic Pump",                                "chung_loai":"-",          "vat_lieu":"-",     "tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "liquid ring pump":                {"mo_ta":"Liquid Ring Pump",                                "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"-",  "don_vi":"pcs","section":"sanitary"},
    "lrp":                             {"mo_ta":"Liquid Ring Pump",                                "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"-",  "don_vi":"pcs","section":"sanitary"},
    "vane pump":                       {"mo_ta":"Vane Pump",                                       "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"-",  "don_vi":"pcs","section":"sanitary"},
    "vp":                              {"mo_ta":"Vane Pump",                                       "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"-",  "don_vi":"pcs","section":"sanitary"},
    "vacuum pump":                     {"mo_ta":"Vacuum Pump",                                     "chung_loai":"-",          "vat_lieu":"-",     "tieu_chuan":"-",  "don_vi":"pcs","section":"sanitary"},
    "vac":                             {"mo_ta":"Vacuum Pump",                                     "chung_loai":"-",          "vat_lieu":"-",     "tieu_chuan":"-",  "don_vi":"pcs","section":"sanitary"},
    "compressor":                      {"mo_ta":"Compressor",                                      "chung_loai":"-",          "vat_lieu":"-",     "tieu_chuan":"-",  "don_vi":"pcs","section":"sanitary"},
    "comp":                            {"mo_ta":"Compressor",                                      "chung_loai":"-",          "vat_lieu":"-",     "tieu_chuan":"-",  "don_vi":"pcs","section":"sanitary"},
    "blower":                          {"mo_ta":"Blower / Fan",                                    "chung_loai":"-",          "vat_lieu":"-",     "tieu_chuan":"-",  "don_vi":"pcs","section":"sanitary"},
    "blw":                             {"mo_ta":"Blower / Fan",                                    "chung_loai":"-",          "vat_lieu":"-",     "tieu_chuan":"-",  "don_vi":"pcs","section":"sanitary"},

    # ── UTILITY VALVE MANUAL ───────────────────────────────────────
    "globe valve manual":              {"mo_ta":"Globe Valve – Manual",                            "chung_loai":"Manual",     "vat_lieu":"Cast Iron","tieu_chuan":"Thread end","don_vi":"pcs","section":"steam"},
    "gv m":                            {"mo_ta":"Globe Valve – Manual",                            "chung_loai":"Manual",     "vat_lieu":"Cast Iron","tieu_chuan":"Thread end","don_vi":"pcs","section":"steam"},
    "ball valve manual":               {"mo_ta":"Ball Valve – Manual",                             "chung_loai":"Manual",     "vat_lieu":"Brass",    "tieu_chuan":"Thread end","don_vi":"pcs","section":"ice_water"},
    "bv m":                            {"mo_ta":"Ball Valve – Manual",                             "chung_loai":"Manual",     "vat_lieu":"Brass",    "tieu_chuan":"Thread end","don_vi":"pcs","section":"ice_water"},
    "butterfly valve manual utility":  {"mo_ta":"Butterfly Valve (Utility) – Manual",              "chung_loai":"Manual",     "vat_lieu":"Cast Iron / Disc SS304","tieu_chuan":"JIS10K","don_vi":"pcs","section":"ice_water"},
    "bfv um":                          {"mo_ta":"Butterfly Valve (Utility) – Manual",              "chung_loai":"Manual",     "vat_lieu":"Cast Iron / Disc SS304","tieu_chuan":"JIS10K","don_vi":"pcs","section":"ice_water"},
    "diaphragm valve manual":          {"mo_ta":"Diaphragm Valve – Manual",                       "chung_loai":"Manual",     "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "dv m":                            {"mo_ta":"Diaphragm Valve – Manual",                       "chung_loai":"Manual",     "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "angle valve manual":              {"mo_ta":"Angle Valve – Manual",                            "chung_loai":"Manual",     "vat_lieu":"SS304",   "tieu_chuan":"Thread end","don_vi":"pcs","section":"air"},
    "av m":                            {"mo_ta":"Angle Valve – Manual",                            "chung_loai":"Manual",     "vat_lieu":"SS304",   "tieu_chuan":"Thread end","don_vi":"pcs","section":"air"},

    # ── UTILITY VALVE ACTUATOR ─────────────────────────────────────
    "globe valve actuator":            {"mo_ta":"Globe Valve – Actuator NC",                       "chung_loai":"Actuator NC","vat_lieu":"Cast Iron","tieu_chuan":"Thread end","don_vi":"pcs","section":"steam"},
    "gv a":                            {"mo_ta":"Globe Valve – Actuator NC",                       "chung_loai":"Actuator NC","vat_lieu":"Cast Iron","tieu_chuan":"Thread end","don_vi":"pcs","section":"steam"},
    "ball valve actuator":             {"mo_ta":"Ball Valve – Actuator NC",                        "chung_loai":"Actuator NC","vat_lieu":"Brass",    "tieu_chuan":"Thread end","don_vi":"pcs","section":"ice_water"},
    "bv a":                            {"mo_ta":"Ball Valve – Actuator NC",                        "chung_loai":"Actuator NC","vat_lieu":"Brass",    "tieu_chuan":"Thread end","don_vi":"pcs","section":"ice_water"},
    "butterfly valve actuator utility":{"mo_ta":"Butterfly Valve (Utility) – Actuator NC",         "chung_loai":"Actuator NC","vat_lieu":"Cast Iron / Disc SS304","tieu_chuan":"JIS10K","don_vi":"pcs","section":"ice_water"},
    "bfv ua":                          {"mo_ta":"Butterfly Valve (Utility) – Actuator NC",         "chung_loai":"Actuator NC","vat_lieu":"Cast Iron / Disc SS304","tieu_chuan":"JIS10K","don_vi":"pcs","section":"ice_water"},
    "angle seat valve actuator":       {"mo_ta":"Angle Seat Valve – Actuator NC",                  "chung_loai":"Actuator NC","vat_lieu":"SS304",   "tieu_chuan":"Thread end","don_vi":"pcs","section":"steam"},
    "asv a":                           {"mo_ta":"Angle Seat Valve – Actuator NC",                  "chung_loai":"Actuator NC","vat_lieu":"SS304",   "tieu_chuan":"Thread end","don_vi":"pcs","section":"steam"},
    "diaphragm valve actuator":        {"mo_ta":"Diaphragm Valve – Actuator NC",                   "chung_loai":"Actuator NC","vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "dv a":                            {"mo_ta":"Diaphragm Valve – Actuator NC",                   "chung_loai":"Actuator NC","vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},

    # ── UTILITY VALVE SOLENOID ─────────────────────────────────────
    "globe valve solenoid":            {"mo_ta":"Globe Valve – Solenoid",                          "chung_loai":"Solenoid",   "vat_lieu":"Cast Iron","tieu_chuan":"Thread end","don_vi":"pcs","section":"steam"},
    "ball valve solenoid":             {"mo_ta":"Ball Valve – Solenoid",                           "chung_loai":"Solenoid",   "vat_lieu":"Brass",    "tieu_chuan":"Thread end","don_vi":"pcs","section":"ice_water"},
    "angle seat valve solenoid":       {"mo_ta":"Angle Seat Valve – Solenoid",                     "chung_loai":"Solenoid",   "vat_lieu":"SS304",   "tieu_chuan":"Thread end","don_vi":"pcs","section":"steam"},
    "asv s":                           {"mo_ta":"Angle Seat Valve – Solenoid",                     "chung_loai":"Solenoid",   "vat_lieu":"SS304",   "tieu_chuan":"Thread end","don_vi":"pcs","section":"steam"},

    # ── UTILITY VALVE MODULATING ───────────────────────────────────
    "globe valve linear control":      {"mo_ta":"Globe Valve – Linear Control (Modulating)",       "chung_loai":"Pneumatic modulating","vat_lieu":"Cast Iron","tieu_chuan":"PN16","don_vi":"pcs","section":"steam"},
    "butterfly valve linear control":  {"mo_ta":"Butterfly Valve – Linear Control (Modulating)",   "chung_loai":"Pneumatic modulating","vat_lieu":"Cast Iron / Disc SS304","tieu_chuan":"JIS10K","don_vi":"pcs","section":"ice_water"},

    # ── SPECIAL VALVES ─────────────────────────────────────────────
    "check valve":                     {"mo_ta":"Check Valve",                                     "chung_loai":"-",          "vat_lieu":"Cast Iron","tieu_chuan":"PN16","don_vi":"pcs","section":"ice_water"},
    "cv":                              {"mo_ta":"Check Valve",                                     "chung_loai":"-",          "vat_lieu":"Cast Iron","tieu_chuan":"PN16","don_vi":"pcs","section":"ice_water"},
    "constant valve":                  {"mo_ta":"Constant Valve (CSTV)",                           "chung_loai":"-",          "vat_lieu":"SS304",   "tieu_chuan":"Thread end","don_vi":"pcs","section":"air"},
    "cstv":                            {"mo_ta":"Constant Valve (CSTV)",                           "chung_loai":"-",          "vat_lieu":"SS304",   "tieu_chuan":"Thread end","don_vi":"pcs","section":"air"},
    "reducing pressure valve":         {"mo_ta":"Reducing Pressure Valve",                         "chung_loai":"-",          "vat_lieu":"Cast Iron","tieu_chuan":"PN16","don_vi":"pcs","section":"steam"},
    "rpv":                             {"mo_ta":"Reducing Pressure Valve",                         "chung_loai":"-",          "vat_lieu":"Cast Iron","tieu_chuan":"PN16","don_vi":"pcs","section":"steam"},
    "air relief valve":                {"mo_ta":"Air Relief Valve",                                "chung_loai":"-",          "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "arv":                             {"mo_ta":"Air Relief Valve",                                "chung_loai":"-",          "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "vacuum relief valve":             {"mo_ta":"Vacuum Relief Valve",                             "chung_loai":"-",          "vat_lieu":"SS304",   "tieu_chuan":"-","don_vi":"pcs","section":"sanitary"},
    "vrv":                             {"mo_ta":"Vacuum Relief Valve",                             "chung_loai":"-",          "vat_lieu":"SS304",   "tieu_chuan":"-","don_vi":"pcs","section":"sanitary"},
    "pressure relief valve":           {"mo_ta":"Pressure Relief Valve",                           "chung_loai":"-",          "vat_lieu":"Brass",   "tieu_chuan":"Thread end","don_vi":"pcs","section":"steam"},
    "prv":                             {"mo_ta":"Pressure Relief Valve",                           "chung_loai":"-",          "vat_lieu":"Brass",   "tieu_chuan":"Thread end","don_vi":"pcs","section":"steam"},

    # ── UTILITY COMPONENTS ─────────────────────────────────────────
    "steam trap":                      {"mo_ta":"Float Steam Trap",                                "chung_loai":"-",          "vat_lieu":"Cast Iron","tieu_chuan":"PN16","don_vi":"pcs","section":"steam"},
    "st":                              {"mo_ta":"Float Steam Trap",                                "chung_loai":"-",          "vat_lieu":"Cast Iron","tieu_chuan":"PN16","don_vi":"pcs","section":"steam"},
    "y strainer":                      {"mo_ta":"Y Strainer",                                      "chung_loai":"-",          "vat_lieu":"Cast Iron","tieu_chuan":"JIS10K","don_vi":"pcs","section":"ice_water"},
    "ys":                              {"mo_ta":"Y Strainer",                                      "chung_loai":"-",          "vat_lieu":"Cast Iron","tieu_chuan":"JIS10K","don_vi":"pcs","section":"ice_water"},
    "expansion vessel":                {"mo_ta":"Expansion Vessel",                                "chung_loai":"-",          "vat_lieu":"SS304",   "tieu_chuan":"Industry","don_vi":"set","section":"ice_water"},
    "ev":                              {"mo_ta":"Expansion Vessel",                                "chung_loai":"-",          "vat_lieu":"SS304",   "tieu_chuan":"Industry","don_vi":"set","section":"ice_water"},
    "syphon 180":                      {"mo_ta":"Syphon 180°",                                     "chung_loai":"-",          "vat_lieu":"SS304",   "tieu_chuan":"Industry","don_vi":"pcs","section":"steam"},
    "sy180":                           {"mo_ta":"Syphon 180°",                                     "chung_loai":"-",          "vat_lieu":"SS304",   "tieu_chuan":"Industry","don_vi":"pcs","section":"steam"},
    "syphon 90":                       {"mo_ta":"Syphon 90°",                                      "chung_loai":"-",          "vat_lieu":"SS304",   "tieu_chuan":"Industry","don_vi":"pcs","section":"steam"},
    "sy90":                            {"mo_ta":"Syphon 90°",                                      "chung_loai":"-",          "vat_lieu":"SS304",   "tieu_chuan":"Industry","don_vi":"pcs","section":"steam"},
    "expansion joint":                 {"mo_ta":"Expansion Joint",                                 "chung_loai":"-",          "vat_lieu":"-",       "tieu_chuan":"-","don_vi":"pcs","section":"ice_water"},
    "ej":                              {"mo_ta":"Expansion Joint",                                 "chung_loai":"-",          "vat_lieu":"-",       "tieu_chuan":"-","don_vi":"pcs","section":"ice_water"},

    # ── INSTRUMENTS ────────────────────────────────────────────────
    "flow meter":                      {"mo_ta":"Flow Meter",                                      "chung_loai":"-",          "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"pcs","section":"instrument"},
    "fm":                              {"mo_ta":"Flow Meter",                                      "chung_loai":"-",          "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"pcs","section":"instrument"},
    "temperature transmitter":         {"mo_ta":"Temperature Transmitter",                         "chung_loai":"-",          "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"set","section":"instrument"},
    "tt":                              {"mo_ta":"Temperature Transmitter",                         "chung_loai":"-",          "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"set","section":"instrument"},
    "temperature indicator":           {"mo_ta":"Temperature Indicator",                           "chung_loai":"-",          "vat_lieu":"SS304",   "tieu_chuan":"SMS","don_vi":"set","section":"instrument"},
    "ti":                              {"mo_ta":"Temperature Indicator",                           "chung_loai":"-",          "vat_lieu":"SS304",   "tieu_chuan":"SMS","don_vi":"set","section":"instrument"},
    "pressure transmitter":            {"mo_ta":"Pressure Transmitter",                            "chung_loai":"-",          "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"set","section":"instrument"},
    "pt":                              {"mo_ta":"Pressure Transmitter",                            "chung_loai":"-",          "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"set","section":"instrument"},
    "pressure indicator":              {"mo_ta":"Pressure Indicator / Gauge",                      "chung_loai":"-",          "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"pcs","section":"instrument"},
    "pi":                              {"mo_ta":"Pressure Indicator / Gauge",                      "chung_loai":"-",          "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"pcs","section":"instrument"},
    "level switch":                    {"mo_ta":"Level Switch",                                    "chung_loai":"-",          "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"set","section":"instrument"},
    "ls":                              {"mo_ta":"Level Switch",                                    "chung_loai":"-",          "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"set","section":"instrument"},
    "conductivity sensor":             {"mo_ta":"Conductivity Sensor",                             "chung_loai":"-",          "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"set","section":"instrument"},
    "cs":                              {"mo_ta":"Conductivity Sensor",                             "chung_loai":"-",          "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"set","section":"instrument"},
    "flow switch":                     {"mo_ta":"Flow Switch",                                     "chung_loai":"-",          "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"set","section":"instrument"},
    "fs":                              {"mo_ta":"Flow Switch",                                     "chung_loai":"-",          "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"set","section":"instrument"},
    "proximity switch":                {"mo_ta":"Proximity Switch",                                "chung_loai":"-",          "vat_lieu":"SS304",   "tieu_chuan":"-",  "don_vi":"pcs","section":"instrument"},
    "ph transmitter":                  {"mo_ta":"pH Transmitter",                                  "chung_loai":"-",          "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"set","section":"instrument"},
    "brix transmitter":                {"mo_ta":"Brix Transmitter",                                "chung_loai":"-",          "vat_lieu":"SS316L",  "tieu_chuan":"SMS","don_vi":"set","section":"instrument"},
}

# ── TETRA PAK LIBRARY — inherits SPX base, overrides/adds TP-specific ─
TPK_OVERRIDES = {
    # Leakage Valve (TP only)
    "leakage valve nc actuator":       {"mo_ta":"Leakage Valve NC – Actuator",                    "chung_loai":"Actuator NC","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "leakage valve no actuator":       {"mo_ta":"Leakage Valve NO – Actuator",                    "chung_loai":"Actuator NO","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "lv nc":                           {"mo_ta":"Leakage Valve NC – Actuator",                    "chung_loai":"Actuator NC","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "lv no":                           {"mo_ta":"Leakage Valve NO – Actuator",                    "chung_loai":"Actuator NO","vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    # Pumps TP
    "centrifugal pump":                {"mo_ta":"Centrifugal Pump",                               "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "cp":                              {"mo_ta":"Centrifugal Pump",                               "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "lobe pump":                       {"mo_ta":"Lobe Pump",                                      "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "lp":                              {"mo_ta":"Lobe Pump",                                      "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "screw pump":                      {"mo_ta":"Screw Pump",                                     "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "sp":                              {"mo_ta":"Screw Pump",                                     "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "utility pump":                    {"mo_ta":"Utility Pump",                                   "chung_loai":"-",          "vat_lieu":"SS304", "tieu_chuan":"Industry","don_vi":"pcs","section":"sanitary"},
    "up":                              {"mo_ta":"Utility Pump",                                   "chung_loai":"-",          "vat_lieu":"SS304", "tieu_chuan":"Industry","don_vi":"pcs","section":"sanitary"},
    "diaphragm pump":                  {"mo_ta":"Diaphragm Pump",                                 "chung_loai":"-",          "vat_lieu":"PP",    "tieu_chuan":"-","don_vi":"pcs","section":"sanitary"},
    "dp":                              {"mo_ta":"Diaphragm Pump",                                 "chung_loai":"-",          "vat_lieu":"PP",    "tieu_chuan":"-","don_vi":"pcs","section":"sanitary"},
    # Sampling (TP names)
    "aseptic sampling valve":          {"mo_ta":"Aseptic Sampling Valve",                         "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "asv":                             {"mo_ta":"Aseptic Sampling Valve",                         "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "sanitary sampling valve":         {"mo_ta":"Sanitary Sampling Valve",                        "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv samp":                        {"mo_ta":"Sanitary Sampling Valve",                        "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    # Filters (TP names)
    "bag filter":                      {"mo_ta":"Bag Filter",                                     "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "bf":                              {"mo_ta":"Bag Filter",                                     "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "inline filter":                   {"mo_ta":"Inline Filter",                                  "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ilf":                             {"mo_ta":"Inline Filter",                                  "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "angle filter":                    {"mo_ta":"Angle Filter",                                   "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "af":                              {"mo_ta":"Angle Filter",                                   "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    # SSV TP (no SV4x codes)
    "ssv 200":                         {"mo_ta":"Single Seat Valve 200",                          "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 210":                         {"mo_ta":"Single Seat Valve 210",                          "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 220":                         {"mo_ta":"Single Seat Valve 220",                          "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
    "ssv 300":                         {"mo_ta":"Single Seat Valve 300",                          "chung_loai":"-",          "vat_lieu":"SS316L","tieu_chuan":"SMS","don_vi":"pcs","section":"sanitary"},
}

# Build TP library = SPX base + overrides
TPK_LIBRARY = {**SPX_LIBRARY, **TPK_OVERRIDES}

SECTION_LABELS = {
    "sanitary":  "ĐƯỜNG PRODUCT / CIP / PROCESS WATER (VI SINH)",
    "ice_water": "ĐƯỜNG ICE WATER / COOLING WATER",
    "steam":     "ĐƯỜNG STEAM / CONDENSATE",
    "air":       "ĐƯỜNG KHÍ NÉN / NƯỚC THÀNH PHỐ",
    "instrument":"INSTRUMENT",
    "equipment": "THIẾT BỊ",
    "unknown":   "THIẾT BỊ CHƯA XÁC ĐỊNH",
}
SECTION_ORDER = ["sanitary", "ice_water", "steam", "air", "instrument", "equipment", "unknown"]
# fmt: on


# ═══════════════════════════════════════════════════════════════════
#  MATCHING ENGINE
# ═══════════════════════════════════════════════════════════════════

def normalize(name: str) -> set:
    """Lowercase + replace -_ with space → set of words."""
    name = name.lower()
    name = re.sub(r"[-_]", " ", name)
    return set(name.split())


def match_library(block_name: str, library: dict):
    """Return (canonical_key, info_dict) or (None, None)."""
    b = normalize(block_name)
    # Exact word-set match first
    for lib_key, info in library.items():
        l = set(lib_key.split())
        if b == l or b.issubset(l) or l.issubset(b):
            return lib_key, info
    return None, None


def should_ignore(block_name: str) -> bool:
    words = normalize(block_name)
    for ign in IGNORE_BLOCKS:
        if set(ign.split()).issubset(words):
            return True
    return False


# ═══════════════════════════════════════════════════════════════════
#  DXF READER
# ═══════════════════════════════════════════════════════════════════

def parse_dxf(dxf_bytes: bytes, library: dict):
    """
    Returns:
        rows  : list of dict  → one per unique (block_name, dn) combination
        unknowns : list of str  → block names not in library
        size_texts : list of {'dn', 'x', 'y'}
        warnings : list of str
    """
    with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as tmp:
        tmp.write(dxf_bytes)
        tmp_path = tmp.name

    doc = ezdxf.readfile(tmp_path)
    msp = doc.modelspace()

    equipments = []  # {'name': str, 'x': float, 'y': float}
    size_texts = []  # {'dn': str, 'x': float, 'y': float}

    for e in msp:
        dtype = e.dxftype()
        if dtype == "INSERT":
            try:
                bname = e.dxf.name
                if should_ignore(bname):
                    continue
                pos = e.dxf.insert
                equipments.append({"name": bname, "x": pos.x, "y": pos.y})
            except Exception:
                pass
        elif dtype in ("TEXT", "MTEXT"):
            try:
                raw = e.text if dtype == "MTEXT" else e.dxf.text
                # Strip MTEXT formatting codes
                raw = re.sub(r"\\[a-zA-Z][^;]*;", "", raw)
                raw = re.sub(r"\{[^}]*\}", "", raw)
                pos = e.dxf.insert
                # Match DN notation
                dn_m = re.search(r"DN\s*\d+", raw)
                if dn_m:
                    dn_str = dn_m.group().replace(" ", "")
                    size_texts.append({"dn": dn_str, "x": pos.x, "y": pos.y})
                # Match inch notation like 1", 1.5", 2"
                inch_m = re.search(r'(\d+(?:\.\d+)?)"', raw)
                if inch_m and not dn_m:
                    size_texts.append({"dn": inch_m.group(), "x": pos.x, "y": pos.y})
            except Exception:
                pass

    # ── Proximity detection ──────────────────────────────────────
    def nearest_dn(ex, ey):
        best, best_d = None, float("inf")
        for t in size_texts:
            d = math.hypot(t["x"] - ex, t["y"] - ey)
            if d < best_d:
                best_d, best = d, t
        if best and best_d <= MAX_DIST:
            return best["dn"], round(best_d, 1)
        return "?", None

    # ── Aggregate by (block_name, dn) ────────────────────────────
    counter = defaultdict(lambda: defaultdict(int))  # block_name → dn → count
    for eq in equipments:
        dn, _ = nearest_dn(eq["x"], eq["y"])
        counter[eq["name"]][dn] += 1

    # ── Build result rows ─────────────────────────────────────────
    rows = []
    unknowns = []
    warnings = []

    for bname, dn_counts in sorted(counter.items()):
        _, info = match_library(bname, library)
        if info is None:
            # Unknown block
            unknowns.append(bname)
            for dn, qty in dn_counts.items():
                rows.append({
                    "block_name": bname,
                    "mo_ta":      f"UNKNOWN — {bname}",
                    "chung_loai": "-",
                    "vat_lieu":   "-",
                    "kt1":        dn,
                    "kt2":        "-",
                    "tieu_chuan": "-",
                    "xuat_xu":    "-",
                    "don_vi":     "pcs",
                    "so_luong":   qty,
                    "section":    "unknown",
                })
        else:
            for dn, qty in dn_counts.items():
                if dn == "?":
                    warnings.append(f"⚠️ {bname}: size chưa xác định (×{qty}) — kiểm tra bản vẽ")
                rows.append({
                    "block_name": bname,
                    "mo_ta":      info["mo_ta"],
                    "chung_loai": info["chung_loai"],
                    "vat_lieu":   info["vat_lieu"],
                    "kt1":        dn,
                    "kt2":        "-",
                    "tieu_chuan": info["tieu_chuan"],
                    "xuat_xu":    "-",
                    "don_vi":     info["don_vi"],
                    "so_luong":   qty,
                    "section":    info["section"],
                })

    return rows, unknowns, size_texts, warnings


# ═══════════════════════════════════════════════════════════════════
#  XLSX BOM BUILDER
# ═══════════════════════════════════════════════════════════════════

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def apply_cell(cell, value, bg_fill, font_color, bold, size=11,
               h_align="center", wrap=True):
    cell.value = value
    cell.fill  = bg_fill
    cell.font  = Font(name="Arial", bold=bold, size=size, color=font_color)
    cell.border = thin_border()
    cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=wrap)


def build_bom_xlsx(rows: list, skill_name: str, project_name: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOQ"

    # ── Column widths ─────────────────────────────────────────────
    col_widths = [5.57, 47.71, 15.14, 16.14, 11.29, 12.00, 10.71, 15.43, 6.71, 9.71]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # ── HEADER COMPANY (rows 1–15) ───────────────────────────────
    header_data = [
        (1,  "ANHMINH",                                              "A1:J1",  24, True,  "center"),
        (2,  "TECHNOLOGY TRADING COMPANY LIMITED",                   "A2:J2",  12, True,  "center"),
        (3,  "Lô C3-7A, Đường số 9, KCN Tân Bình, TP. HCM",         "A3:J3",  11, False, "center"),
        (4,  "Tel: (028) 2246 6688 | Fax: (028) 2246 6699",          "A4:J4",  11, False, "center"),
        (5,  "Email: info@anhminh.com.vn | MST: 0315xxxxxx",         "A5:J5",  11, False, "center"),
        (6,  "FACSIMILE TRANSMISSION",                                "A6:J6",  14, True,  "center"),
    ]
    for row_n, text, merge_range, fsize, bold, align in header_data:
        ws.merge_cells(merge_range)
        c = ws.cell(row=row_n, column=1, value=text)
        c.font = Font(name="Arial", bold=bold, size=fsize, color=BLUE)
        c.alignment = Alignment(horizontal=align, vertical="center")
        ws.row_dimensions[row_n].height = fsize + 6

    # ── Meta rows 7–11 ────────────────────────────────────────────
    from datetime import date
    meta = [
        (7,  f"Ngày / Date: {date.today().strftime('%d/%m/%Y')}"),
        (8,  "Gởi từ / From: ANHMINH"),
        (9,  "Số trang / Pages: 1"),
        (10, f"BG số / No.: BOM-{date.today().strftime('%Y%m%d')}"),
        (11, f"Dự án / Project: {project_name}"),
    ]
    for row_n, text in meta:
        ws.cell(row=row_n, column=1, value=text).font = Font(name="Arial", size=11)
        ws.merge_cells(f"A{row_n}:J{row_n}")
        ws.row_dimensions[row_n].height = 15

    # Row 12: BÁO GIÁ/QUOTATION
    ws.merge_cells("A12:J12")
    c12 = ws.cell(row=12, column=1, value="BÁO GIÁ / QUOTATION")
    c12.font = Font(name="Arial", bold=True, size=18, color=RED)
    c12.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[12].height = 28

    # Rows 13–15
    for row_n, text in [
        (13, f"Kính gửi Quý khách hàng. Chúng tôi xin trân trọng chào giá các vật tư theo đề mục bên dưới — Skill: {skill_name}."),
        (14, "A /  PHẠM VI CÔNG VIỆC / SCOPE OF WORK: Cung cấp vật tư, thiết bị theo bản vẽ P&ID."),
        (15, "B /  ĐƠN GIÁ / PRICE: Theo bảng bên dưới."),
    ]:
        ws.merge_cells(f"A{row_n}:J{row_n}")
        c = ws.cell(row=row_n, column=1, value=text)
        c.font = Font(name="Arial", bold=(row_n in (14, 15)), size=11)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row_n].height = 15 if row_n != 13 else 20

    # ── COLUMN HEADER rows 16–18 ──────────────────────────────────
    col_headers = ["STT", "MÔ TẢ", "CHỦNG LOẠI", "VẬT LIỆU",
                   "K.THƯỚC 1", "K.THƯỚC 2", "TIÊU CHUẨN", "XUẤT XỨ", "ĐƠN VỊ", "SỐ LƯỢNG"]
    for r in range(16, 19):
        for col, hdr in enumerate(col_headers, 1):
            c = ws.cell(row=r, column=col, value=hdr if r == 16 else "")
            c.fill   = BG_CYAN
            c.font   = Font(name="Arial", bold=True, size=11, color=RED)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 15
    ws.merge_cells("A16:A18")
    ws.merge_cells("B16:B18")
    for col in range(3, 11):
        ws.merge_cells(f"{chr(64+col)}16:{chr(64+col)}18")

    # ── DATA rows (19+) ───────────────────────────────────────────
    current_row = 19

    def write_section_row(stt, label, level="major"):
        nonlocal current_row
        bg = BG_CYAN if level in ("major", "sub") else BG_YELLOW
        vals = [stt, label] + [""] * 8
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=current_row, column=col, value=val)
            c.fill   = bg
            c.font   = Font(name="Arial", bold=True, size=11, color=RED)
            c.border = thin_border()
            c.alignment = Alignment(
                horizontal="left" if col == 2 else "center",
                vertical="center", wrap_text=True,
            )
        ws.row_dimensions[current_row].height = 15
        current_row += 1

    def write_data_row(stt, mo_ta, chung_loai, vat_lieu, kt1, kt2,
                       tieu_chuan, xuat_xu, don_vi, so_luong):
        nonlocal current_row
        vals = [stt, mo_ta, chung_loai, vat_lieu, kt1, kt2,
                tieu_chuan, xuat_xu, don_vi, so_luong]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=current_row, column=col, value=val)
            c.fill   = BG_YELLOW
            c.font   = Font(name="Arial", bold=True, size=11, color=RED)
            c.border = thin_border()
            c.alignment = Alignment(
                horizontal="left" if col == 2 else "center",
                vertical="center", wrap_text=True,
            )
        ws.row_dimensions[current_row].height = 15
        current_row += 1

    # ── Group rows by section ─────────────────────────────────────
    grouped = defaultdict(list)
    for r in rows:
        grouped[r["section"]].append(r)

    write_section_row("A", "PHẦN CƠ KHÍ / MECHANICAL PART", level="major")

    stt_roman = ["I", "II", "III", "IV", "V", "VI", "VII"]
    roman_idx = 0

    for sec in SECTION_ORDER:
        if sec not in grouped:
            continue
        sec_rows = grouped[sec]
        if not sec_rows:
            continue

        sec_label = SECTION_LABELS.get(sec, sec.upper())
        write_section_row(stt_roman[roman_idx], sec_label, level="sub")
        roman_idx += 1

        item_idx = 1
        for r in sec_rows:
            write_data_row(
                str(item_idx),
                r["mo_ta"],
                r["chung_loai"],
                r["vat_lieu"],
                r["kt1"],
                r["kt2"],
                r["tieu_chuan"],
                r["xuat_xu"],
                r["don_vi"],
                r["so_luong"],
            )
            item_idx += 1

    # ── FOOTER ───────────────────────────────────────────────────
    for label in ["TOTAL (VND)", "VAT 10% (VND)", "GRAND TOTAL (VND)"]:
        ws.merge_cells(f"A{current_row}:J{current_row}")
        c = ws.cell(row=current_row, column=1, value=label)
        c.font  = Font(name="Arial", bold=True, size=11)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 15
        current_row += 1

    footer_notes = [
        "* Ghi chú / Notes:",
        "- Giá này không bao gồm thuế VAT 10%.",
        "- Không bao gồm chi phí vận chuyển, lắp đặt.",
        "C  ĐIỀU KIỆN BÁN HÀNG / SALES CONDITION:",
        "-  Địa điểm giao hàng / Delivery destination: Tại kho ANHMINH, TP. HCM",
        "-  Thời gian giao hàng / Lead time: 7-10 tuần",
        "-  Điều kiện thanh toán / Payment term: 30% tạm ứng – 40% tập kết – 30% sau khi hoàn thành",
        "-  Thời hạn chào giá / Validity until: 30 ngày kể từ ngày gửi báo giá",
        "Chân thành cảm ơn Quý khách hàng đã tin tưởng và hợp tác.",
        "Trân trọng kính chào / Best regards,",
        "ANHMINH Technology Trading Company Limited",
    ]
    for note in footer_notes:
        ws.merge_cells(f"A{current_row}:J{current_row}")
        c = ws.cell(row=current_row, column=1, value=note)
        bold = note.startswith(("*", "C "))
        color = RED if note.startswith("*") else BLACK
        c.font = Font(name="Arial", bold=bold, size=10, color=color)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[current_row].height = 13
        current_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ═══════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="P&ID BOM Extractor — ANH MINH",
    page_icon="🔧",
    layout="wide",
)

# ── Custom CSS ────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-title {
        font-size: 2.2rem; font-weight: 800;
        color: #003399; margin-bottom: 0;
    }
    .sub-title {
        font-size: 1rem; color: #666; margin-top: 0;
    }
    .stDataFrame thead th {background-color:#00FFFF; color:#FF0000; font-weight:bold;}
    .section-header {
        background: #00FFFF; color: #FF0000;
        font-weight: 700; padding: 6px 12px;
        border-radius: 4px; margin: 12px 0 4px 0;
    }
    .unknown-box {
        background: #FFF3CD; border-left: 4px solid #FFA500;
        padding: 8px 12px; border-radius: 4px; margin-top: 8px;
        font-size: 0.88rem;
    }
    .warn-box {
        background: #F8D7DA; border-left: 4px solid #DC3545;
        padding: 8px 12px; border-radius: 4px; margin-top: 8px;
        font-size: 0.88rem;
    }
    .success-box {
        background: #D4EDDA; border-left: 4px solid #28A745;
        padding: 8px 12px; border-radius: 4px; margin-top: 8px;
        font-size: 0.88rem;
    }
</style>
""", unsafe_allow_html=True)

# ── TITLE ─────────────────────────────────────────────────────────
st.markdown('<div class="main-title">🔧 P&ID BOM Extractor</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">ANH MINH 2021 Standard — SPX / Tetra Pak | Rule-based · No AI cost</div>',
            unsafe_allow_html=True)
st.markdown("---")

# ── SIDEBAR CONFIG ────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Cấu hình")
    skill_choice = st.radio(
        "Chọn tiêu chuẩn P&ID:",
        ["SPX Type (ANH MINH 2021)", "Tetra Pak Type (ANH MINH 2021)"],
        help="SPX dùng SV41/43/44/42. Tetra Pak có thêm Leakage Valve, Screw Pump, Angle Filter.",
    )
    project_name = st.text_input("Tên dự án", value="ANHMINH Project 2025")
    st.markdown("---")
    st.markdown("**Quy tắc xác định size:**")
    st.markdown(f"• Tìm chú thích DN gần thiết bị nhất  \n• Ngưỡng tối đa: **{MAX_DIST} đơn vị DXF**  \n• Nếu không tìm thấy → ghi `?`")
    st.markdown("---")
    st.markdown("**Quy tắc nhận diện thiết bị:**")
    st.markdown("• Đọc **Block Name** từ DXF  \n• So khớp không phân biệt thứ tự từ  \n• Không đoán, không dùng AI")

library_map = {
    "SPX Type (ANH MINH 2021)": SPX_LIBRARY,
    "Tetra Pak Type (ANH MINH 2021)": TPK_LIBRARY,
}
selected_library = library_map[skill_choice]

# ── FILE UPLOAD ───────────────────────────────────────────────────
uploaded = st.file_uploader(
    "📂 Upload file DXF (P&ID bản vẽ)",
    type=["dxf"],
    help="Chỉ hỗ trợ định dạng .dxf. File sẽ được xử lý hoàn toàn cục bộ.",
)

if uploaded is None:
    st.info("⬆️  Vui lòng upload file DXF để bắt đầu trích xuất BOM.")
    st.stop()

# ── PARSE DXF ─────────────────────────────────────────────────────
with st.spinner("🔍 Đang phân tích file DXF…"):
    try:
        dxf_bytes = uploaded.read()
        rows, unknowns, size_texts, warnings = parse_dxf(dxf_bytes, selected_library)
    except Exception as exc:
        st.error(f"❌ Lỗi đọc file DXF: {exc}")
        st.stop()

total_eq   = sum(r["so_luong"] for r in rows)
total_types = len(set(r["block_name"] for r in rows))

# ── SUMMARY METRICS ───────────────────────────────────────────────
col1, col2, col3, col4 = st.columns(4)
col1.metric("📦 Tổng thiết bị", total_eq)
col2.metric("🔢 Loại block", total_types)
col3.metric("📐 Chú thích size", len(size_texts))
col4.metric("❓ Block chưa nhận diện", len(unknowns))

st.markdown("---")

# ── WARNINGS ─────────────────────────────────────────────────────
if warnings:
    with st.expander(f"⚠️ {len(warnings)} cảnh báo size chưa xác định", expanded=False):
        for w in warnings:
            st.markdown(f'<div class="warn-box">{w}</div>', unsafe_allow_html=True)

if unknowns:
    with st.expander(f"❓ {len(unknowns)} block name chưa có trong thư viện ({skill_choice})", expanded=False):
        for u in unknowns:
            st.markdown(f'<div class="unknown-box">🔷 <code>{u}</code></div>', unsafe_allow_html=True)
        st.caption("Các block này vẫn được xuất vào BOM với mô tả 'UNKNOWN'. Hãy kiểm tra và thêm vào thư viện nếu cần.")

# ── BOM PREVIEW ───────────────────────────────────────────────────
st.subheader("📋 Xem trước BOM")

# Group by section
grouped = defaultdict(list)
for r in rows:
    grouped[r["section"]].append(r)

all_preview_rows = []
for sec in SECTION_ORDER:
    if sec not in grouped:
        continue
    sec_label = SECTION_LABELS.get(sec, sec.upper())
    st.markdown(f'<div class="section-header">▸ {sec_label}</div>', unsafe_allow_html=True)
    sec_data = grouped[sec]
    df_sec = pd.DataFrame([{
        "STT":        i + 1,
        "Mô tả":      r["mo_ta"],
        "Chủng loại": r["chung_loai"],
        "Vật liệu":   r["vat_lieu"],
        "K.Thước":    r["kt1"],
        "Tiêu chuẩn": r["tieu_chuan"],
        "Đơn vị":     r["don_vi"],
        "Số lượng":   r["so_luong"],
    } for i, r in enumerate(sec_data)])
    st.dataframe(df_sec, use_container_width=True, hide_index=True)
    all_preview_rows.extend(sec_data)

# ── VERIFICATION SUMMARY ──────────────────────────────────────────
st.markdown("---")
passed = all(r["kt1"] != "?" for r in rows if r["section"] != "unknown")
if passed and not unknowns:
    st.markdown('<div class="success-box">✅ BOM đã kiểm tra: tất cả thiết bị đã có size — PASS</div>',
                unsafe_allow_html=True)
elif passed:
    st.markdown(f'<div class="unknown-box">⚠️ BOM PASS về size, nhưng có {len(unknowns)} block chưa nhận diện.</div>',
                unsafe_allow_html=True)
else:
    st.markdown('<div class="warn-box">⚠️ Một số thiết bị chưa xác định được size (size = ?). Kiểm tra lại bản vẽ.</div>',
                unsafe_allow_html=True)

# ── EXPORT ───────────────────────────────────────────────────────
st.markdown("---")
st.subheader("📥 Xuất BOM")

col_dl1, col_dl2 = st.columns([1, 3])
with col_dl1:
    if st.button("⚙️ Tạo file BOM Excel", type="primary", use_container_width=True):
        with st.spinner("Đang tạo file xlsx…"):
            xlsx_bytes = build_bom_xlsx(rows, skill_choice, project_name)
        st.session_state["xlsx_bytes"] = xlsx_bytes
        st.success("✅ File đã sẵn sàng!")

if "xlsx_bytes" in st.session_state:
    from datetime import date
    fname = f"BOM_{project_name.replace(' ','_')}_{date.today().strftime('%Y%m%d')}.xlsx"
    col_dl2.download_button(
        label="⬇️  Tải xuống BOM.xlsx",
        data=st.session_state["xlsx_bytes"],
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

st.caption("ANHMINH 2021 Standard | P&ID BOM Extractor v1.0 | Không tốn API cost — 100% rule-based")
