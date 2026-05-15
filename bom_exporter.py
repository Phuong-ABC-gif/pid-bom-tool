"""
bom_exporter.py  —  ANH MINH 2021
Xuất BOM Excel chuẩn FORM_BOM_STANDARD.

Màu sắc:
  Major section (A/B/C) : nền CYAN   #00FFFF | chữ đỏ | bold
  Sub-section  (I/II/III): nền VÀNG  #FFFF00 | chữ đỏ | bold
  Data rows (thiết bị)  : KHÔNG nền  | chữ đen | KHÔNG bold

Sắp xếp trong mỗi sub-section (line):
  Cấp 1: display_name A→Z   (gom cùng loại thiết bị)
  Cấp 2: size giảm dần       (DN100 → DN65 → DN50 → DN25 ...)
  Cấp 3: chung_loai A→Z     (tie-break)
"""
import io
import re
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Màu ───────────────────────────────────────────────────────────────────
BG_CYAN   = PatternFill("solid", start_color="00FFFF", end_color="00FFFF")
BG_YELLOW = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")
BG_NONE   = PatternFill(fill_type=None)
FG_RED    = "FF0000"
FG_BLACK  = "000000"

COL_WIDTHS  = [5.57, 47.71, 15.14, 16.14, 11.29, 12.00, 10.71, 15.43, 6.71, 9.71]
COL_HEADERS = ["STT", "MÔ TẢ", "CHỦNG LOẠI", "VẬT LIỆU",
               "K.THƯỚC 1", "K.THƯỚC 2", "TIÊU CHUẨN", "XUẤT XỨ", "ĐƠN VỊ", "SỐ LƯỢNG"]

ROMAN = ["I","II","III","IV","V","VI","VII","VIII","IX","X",
         "XI","XII","XIII","XIV","XV","XVI","XVII","XVIII","XIX","XX"]

# Thứ tự xuất hiện trong BOM
SECTION_ORDER = [
    ("Product",          "ĐƯỜNG PRODUCT (SANITARY)"),
    ("CIP",              "ĐƯỜNG CIP (SANITARY)"),
    ("RO/Process Water", "ĐƯỜNG RO / PROCESS WATER (SANITARY)"),
    ("Ice Water",        "ĐƯỜNG ICE WATER (CHILLER)"),
    ("Cooling Water",    "ĐƯỜNG COOLING WATER"),
    ("Steam/Condensate", "ĐƯỜNG STEAM & CONDENSATE"),
    ("Compressed Air",   "ĐƯỜNG COMPRESSED AIR"),
    ("City/Soft Water",  "ĐƯỜNG CITY WATER / SOFT WATER"),
    ("Unknown",          "CHƯA XÁC ĐỊNH LINE"),
]


# ══════════════════════════════════════════════════════════════════════════════
# SORT HELPER
# ══════════════════════════════════════════════════════════════════════════════

def _size_to_mm(size_str: str) -> float:
    """
    Chuyển chuỗi size → mm để so sánh (lớn hơn = số lớn hơn).
    DN50  → 50.0
    2"    → 50.8   (2 × 25.4)
    1.5"  → 38.1
    1/2"  → 12.7
    1 1/2"→ 38.1
    ?     → -1     (đặt cuối)
    """
    s = str(size_str).strip()
    if not s or s == "?":
        return -1.0

    # DN format: DN50, DN 50, DN50.8
    m = re.match(r"DN\s*(\d+(?:\.\d+)?)", s, re.IGNORECASE)
    if m:
        return float(m.group(1))

    # Inch compound: "1 1/2\"" → 1.5 * 25.4
    m = re.match(r"(\d+)\s+(\d+)/(\d+)", s)
    if m:
        whole = int(m.group(1))
        num   = int(m.group(2))
        den   = int(m.group(3))
        return (whole + num / den) * 25.4

    # Fraction: "1/2\""
    m = re.match(r"(\d+)/(\d+)", s)
    if m:
        return int(m.group(1)) / int(m.group(2)) * 25.4

    # Decimal inch: "1.5\""
    m = re.search(r"(\d+(?:\.\d+)?)", s)
    if m:
        return float(m.group(1)) * 25.4

    return -1.0


def _row_sort_key(row: dict) -> tuple:
    """
    Sort key cho 1 data row trong 1 section (line_type):
      (display_name asc, size desc, chung_loai asc)
    """
    return (
        row["display_name"].lower(),   # A→Z
        -_size_to_mm(row["size"]),     # lớn → nhỏ  (negate để sort asc = desc thực)
        row["chung_loai"].lower(),     # tie-break
    )


# ══════════════════════════════════════════════════════════════════════════════
# AGGREGATE  (gom + đếm số lượng, giữ nguyên mỗi combination riêng biệt)
# ══════════════════════════════════════════════════════════════════════════════

def _aggregate(items: list) -> dict:
    """
    items → dict {line_type: [row_dict, ...]}  đã sắp xếp đúng thứ tự.

    Mỗi row_dict duy nhất được xác định bởi:
        (display_name, size, line_type, chung_loai, vat_lieu, tieu_chuan, don_vi)
    → đếm số lượng (so_luong)

    Sau đó sort mỗi nhóm line_type theo _row_sort_key.
    """
    # Bước 1: đếm số lượng theo combination
    counts: dict = {}
    for item in items:
        key = (
            item.get("display_name", ""),
            item.get("size", "?"),
            item.get("line_type", "Unknown"),
            item.get("chung_loai", "-"),
            item.get("vat_lieu", "-"),
            item.get("tieu_chuan", "-"),
            item.get("don_vi", "pcs"),
            item.get("line_label", ""),
        )
        counts[key] = counts.get(key, 0) + 1

    # Bước 2: chuyển sang list of dict
    all_rows = []
    for key, cnt in counts.items():
        dn, sz, lt, cl, vl, tc, dv, lbl = key
        all_rows.append({
            "display_name": dn,
            "size":         sz,
            "line_type":    lt,
            "line_label":   lbl,
            "chung_loai":   cl,
            "vat_lieu":     vl,
            "tieu_chuan":   tc,
            "don_vi":       dv,
            "so_luong":     cnt,
        })

    # Bước 3: nhóm theo line_type, sort mỗi nhóm
    grouped: dict = {}
    for row in all_rows:
        lt = row["line_type"]
        grouped.setdefault(lt, []).append(row)

    # *** SORT — cấp 1 tên A→Z, cấp 2 size lớn→nhỏ, cấp 3 chung_loai A→Z ***
    for lt in grouped:
        grouped[lt] = sorted(grouped[lt], key=_row_sort_key)

    return grouped


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, value, bg, color, bold, h_align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = bg
    c.font      = Font(name="Arial", bold=bold, size=11, color=color)
    c.border    = _thin()
    c.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)
    return c


def _write_row(ws, row_idx: int, values: list, bg, color, bold) -> int:
    for col, val in enumerate(values, 1):
        halign = "left" if col == 2 else "center"
        _cell(ws, row_idx, col, val, bg, color, bold, halign)
    ws.row_dimensions[row_idx].height = 16
    return row_idx + 1


def write_major(ws, r, stt, label):
    return _write_row(ws, r, [stt, label, "", "", "", "", "", "", "SET", 1],
                      BG_CYAN, FG_RED, True)


def write_sub(ws, r, stt, label):
    return _write_row(ws, r, [stt, label, "", "", "", "", "", "", "SET", 1],
                      BG_YELLOW, FG_RED, True)


def write_data(ws, r, stt, mo_ta, chung_loai, vat_lieu,
               kt1, kt2, tieu_chuan, xuat_xu, don_vi, so_luong):
    return _write_row(ws, r,
                      [stt, mo_ta, chung_loai, vat_lieu, kt1, kt2,
                       tieu_chuan, xuat_xu, don_vi, so_luong],
                      BG_NONE, FG_BLACK, False)


# ══════════════════════════════════════════════════════════════════════════════
# HEADER / FOOTER
# ══════════════════════════════════════════════════════════════════════════════

def _build_header(ws) -> int:
    ws.merge_cells("A1:J1")
    ws["A1"].value = "ANHMINH TECHNOLOGY TRADING COMPANY LIMITED"
    ws["A1"].font  = Font(name="Arial", bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:J2")
    ws["A2"].value     = "Address: ..."
    ws["A2"].font      = Font(name="Arial", size=10)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:J3")
    ws["A3"].value     = f"BÁO GIÁ / QUOTATION — {date.today().strftime('%d/%m/%Y')}"
    ws["A3"].font      = Font(name="Arial", bold=True, size=14)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    for r in range(4, 10):
        ws.row_dimensions[r].height = 8

    for r in range(10, 12):
        for col, hdr in enumerate(COL_HEADERS, 1):
            c = ws.cell(row=r, column=col, value=(hdr if r == 10 else None))
            c.fill      = BG_CYAN
            c.font      = Font(name="Arial", bold=True, size=11, color=FG_RED)
            c.border    = _thin()
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 18

    for col in range(1, 11):
        ws.merge_cells(start_row=10, start_column=col, end_row=11, end_column=col)

    for col, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return 12  # data bắt đầu từ row 12


def _build_footer(ws, row: int):
    for label in ["TOTAL (VND)", "VAT 10% (VND)", "GRAND TOTAL (VND)"]:
        ws.merge_cells(f"A{row}:J{row}")
        c = ws[f"A{row}"]
        c.value     = label
        c.font      = Font(name="Arial", bold=True, size=11)
        c.border    = _thin()
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 16
        row += 1

    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"].value = "* Ghi chú / Notes:"
    ws[f"A{row}"].font  = Font(name="Arial", bold=True, size=11, color=FG_RED)
    row += 1
    for note in [
        "- Giá này không bao gồm VAT.",
        "- Thời gian giao hàng / Lead time: 7–10 tuần.",
        "- Điều kiện thanh toán: 30% tạm ứng – 40% tập kết – 30% sau hoàn thành.",
        "- Thời hạn chào giá / Validity: 30 ngày kể từ ngày gửi báo giá.",
        "Trân trọng kính chào / Best regards",
    ]:
        ws.merge_cells(f"A{row}:J{row}")
        ws[f"A{row}"].value = note
        ws[f"A{row}"].font  = Font(name="Arial", size=10)
        row += 1


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC API
# ══════════════════════════════════════════════════════════════════════════════

def export_bom(items: list) -> bytes:
    """
    Nhận list items từ dxf_processor → xuất bytes file .xlsx.
    Mỗi item: dict có các key display_name, size, line_type, line_label,
              chung_loai, vat_lieu, tieu_chuan, don_vi.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"
    ws.sheet_view.showGridLines = False

    r = _build_header(ws)

    # Aggregate + sort
    grouped = _aggregate(items)

    # Viết PHẦN CƠ KHÍ (major header — chỉ 1 lần)
    r = write_major(ws, r, "A", "PHẦN CƠ KHÍ / MECHANICAL")

    roman_idx = 0
    for line_type, section_label in SECTION_ORDER:
        rows = grouped.get(line_type)
        if not rows:
            continue

        # Sub-section header (I, II, III...)
        r = write_sub(ws, r, ROMAN[roman_idx], section_label)
        roman_idx += 1

        # Data rows — đã được sort bởi _aggregate
        for idx, row_data in enumerate(rows, 1):
            r = write_data(
                ws, r,
                stt        = str(idx),
                mo_ta      = row_data["display_name"],
                chung_loai = row_data["chung_loai"],
                vat_lieu   = row_data["vat_lieu"],
                kt1        = row_data["size"],
                kt2        = "-",
                tieu_chuan = row_data["tieu_chuan"],
                xuat_xu    = "",
                don_vi     = row_data["don_vi"],
                so_luong   = row_data["so_luong"],
            )

    _build_footer(ws, r + 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
